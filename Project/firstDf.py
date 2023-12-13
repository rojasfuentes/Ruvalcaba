from list import excel_files, ruta_destino
import openpyxl
import pandas as pd
import os

archivos_con_error = []

columnas = [
    'PROVEEDOR', 'FECHA DE RECEPCIÓN', 'Número de factura', 'Guía de Embarque / Carta Porte',
    'Cliente', 'Número de Nota', 'Fecha', 'Destinatario', 'Localidad', 'Bultos',
    'Peso', 'Tipo de Envió', 'Tipo de Unidad', 'Flete', 'Reparto', 'Km. Adicionales',
    'Maniobras', 'Estadías', 'Retorno', 'Otros', 'Total'
]

# Crear un DataFrame vacío con las columnas definidas
nuevo_df = pd.DataFrame(columns=columnas)
#print(excel_files)

#print(len(excel_files))
#ruta_excel = r'C:\Users\JFROJAS\Desktop\11. NOVIEMBRE\TPE 01 NOVIEMBRE\ANDRADE\Revisión_27_Octubre_2023_Transportes Andrade Y Ortega, S. A. de C.V-SEP.xlsx'

for ruta_excel in excel_files:  # Itera sobre cada ruta de archivo en excel_files
    try:
        workbook = openpyxl.load_workbook(ruta_excel)
        sheet = workbook.active

        workbook = openpyxl.load_workbook(ruta_excel)
        sheet = workbook.active

        # Encontrar la última fila con datos en la columna C
        columna_C = sheet['C']
        ultima_fila = len(columna_C)

        for idx, cell in enumerate(columna_C, start=1):
            if cell.value is None:
                ultima_fila = idx - 1
                break

        # Crear un DataFrame de Pandas con los valores desde C9 hasta V(última fila con datos)
        data = []
        for row in sheet.iter_rows(min_row=9, max_row=ultima_fila, min_col=3, max_col=22, values_only=True):
            data.append(row)

        column_names = [f'Celda_{i}' for i in range(3, 23)]  # Obtener nombres de columna

        df = pd.DataFrame(data, columns=column_names)



        df.columns = df.iloc[0]
        df = df[1:]  # Eliminar la primera fila (ya utilizada como nombres de columna)

        ##########

        # Encuentra índices de filas con dos 'None' consecutivos en la primera columna
        indices = df[df.columns[0]].index[df[df.columns[0]].isnull()]
        for i in range(len(indices) - 1):
            if indices[i + 1] - indices[i] == 1:
                break

        # Elimina las filas a partir del índice donde están los 'None' consecutivos
        if len(indices) > 0:
            for i in range(len(indices) - 1):
                if indices[i + 1] - indices[i] == 1:
                    break

            if i < len(indices) - 1:
                df = df.iloc[:indices[i]]

        valor_celda_E2 = sheet['E2'].value
        valor_celda_E4 = sheet['E4'].value

        # Eliminar la última fila del DataFrame
        df = df.iloc[:-1]


        print(df)


        nuevo_df['PROVEEDOR'] = [valor_celda_E2] * len(nuevo_df)
        nuevo_df['FECHA DE RECEPCIÓN'] = [valor_celda_E4] * len(nuevo_df)
        nuevo_df['Número de factura'] = df['Número de factura']
        nuevo_df['Cliente'] = df['Cliente']
        nuevo_df['Número de Nota'] = df['Número de Nota']
        nuevo_df['Fecha'] = df['Fecha']
        nuevo_df['Destinatario'] = df['Destinatario']
        nuevo_df['Localidad'] = df['Localidad']
        nuevo_df['Bultos'] = df['Bultos']
        nuevo_df['Tipo de Envió\nL=Local,\nF=Foráneo'] = df['Tipo de Envió\nL=Local,\nF=Foráneo']
        nuevo_df['Tipo de Unidad'] = df['Tipo de Unidad']
        nuevo_df['Flete'] = df['Flete']
        nuevo_df['Reparto'] = df['Reparto']
        nuevo_df['Km. Adicionales'] = df['Km. Adicionales']
        nuevo_df['Maniobras'] = df['Maniobras']
        nuevo_df['Estadías'] = df['Estadías']
        nuevo_df['Retorno'] = df['Retorno']
        nuevo_df['Otros'] = df['Otros']
        nuevo_df['Total'] = df['Total']


        print(nuevo_df)

        ruta_guardado = ruta_destino  
        nombre_archivo = 'resultado_' + ruta_excel.split('\\')[-1]  # Obtener el nombre del archivo de la ruta
        #nuevo_df.to_excel(ruta_guardado + nombre_archivo, index=False)  # Guardar el DataFrame en un archivo Excel

        ruta_guardado = os.path.join(ruta_destino, nombre_archivo)
        nuevo_df.to_excel(ruta_guardado, index=False)

        print(f"Archivo '{nombre_archivo}' procesado y guardado con éxito.")
    
    except Exception as e:
        archivos_con_error.append(ruta_excel)
        print(f"Error al procesar el archivo '{ruta_excel}': {str(e)}")
        continue  # Continuar con el siguiente archivo después de capturar la excepción

# Después de terminar el bucle, imprime los archivos con errores
print("Archivos con errores:")
for archivo in archivos_con_error:
    print(archivo)



